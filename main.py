import os
import json
import logging
from datetime import datetime
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from src.analysis.cv_analyzer import CVAnalyzer
from config import (
    CV_DIR, 
    SKILLS_DB, 
    REPORTS_DIR, 
    VALID_EXTENSIONS
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(REPORTS_DIR, 'cv_analysis.log')),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def get_cv_files() -> List[str]:
    """Get list of CV files from the CV directory."""
    cv_files = []
    for file in os.listdir(CV_DIR):
        if any(file.endswith(ext) for ext in VALID_EXTENSIONS):
            cv_files.append(os.path.join(CV_DIR, file))
    return cv_files

def save_results(results: List[Dict[str, Any]]) -> str:
    """Save analysis results to a JSON file."""
    if not results:
        logger.warning("No results to save.")
        return ""

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(REPORTS_DIR, f'analysis_report_{timestamp}.json')
    
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
        logger.info(f"Results saved to {output_file}")
        return output_file
    except Exception as e:
        logger.error(f"Error saving results: {str(e)}")
        return ""

def export_skills_to_excel(summary: Dict[str, Any]) -> str:
    """
    Export CV skills to an Excel file with consolidated skills per CV.
    
    Args:
        summary (dict): Summary dictionary containing CV skills
    
    Returns:
        str: Path to the generated Excel file
    """
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CV Skills Analysis"

    # Set up headers
    headers = ['CV Name', 'Technical Skills', 'Professional Skills', 'Soft Skills', 'Additional Skills']
    ws.append(headers)

    # Style the headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Populate data
    cv_skills = summary.get('cv_skills', {})
    for cv_name, skill_categories in cv_skills.items():
        # Prepare skill lists for each category
        skills_by_category = {
            'Technical Skills': [],
            'Professional Skills': [],
            'Soft Skills': [],
            'Additional Skills': []
        }

        # Categorize skills
        for category, skills in skill_categories.items():
            # Normalize category names
            normalized_category = category.lower()
            
            if 'technical' in normalized_category or 'it' in normalized_category or 'programming' in normalized_category:
                skills_by_category['Technical Skills'].extend(skills)
            elif 'professional' in normalized_category or 'management' in normalized_category:
                skills_by_category['Professional Skills'].extend(skills)
            elif 'soft' in normalized_category or 'communication' in normalized_category or 'interpersonal' in normalized_category:
                skills_by_category['Soft Skills'].extend(skills)
            else:
                skills_by_category['Additional Skills'].extend(skills)

        # Remove duplicates and sort skills
        for category in skills_by_category:
            skills_by_category[category] = sorted(set(skills_by_category[category]))

        # Create row with consolidated skills
        row = [
            cv_name,
            ', '.join(skills_by_category['Technical Skills']) or 'N/A',
            ', '.join(skills_by_category['Professional Skills']) or 'N/A',
            ', '.join(skills_by_category['Soft Skills']) or 'N/A',
            ', '.join(skills_by_category['Additional Skills']) or 'N/A'
        ]
        
        ws.append(row)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_filename = os.path.join(REPORTS_DIR, f'cv_skills_analysis_{timestamp}.xlsx')
    
    # Save the workbook
    wb.save(excel_filename)
    
    logger.info(f"Excel report saved to {excel_filename}")
    return excel_filename

def generate_summary(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Generate a summary of the analysis results."""
    summary = {
        'total_cvs_processed': len(results),
        'cvs_with_skills': 0,
        'most_common_skills': {},
        'cv_skills': {},  # New dictionary to store skills for each CV
        'analysis_timestamp': datetime.now().isoformat()
    }

    # Count skills across all CVs
    skill_count = {}
    for result in results:
        if 'filename' in result and 'skills_found' in result:
            # Track CVs with skills
            summary['cvs_with_skills'] += 1
            
            # Store skills for this specific CV
            summary['cv_skills'][result['filename']] = {}
            
            for category, skills in result.get('skills_found', {}).items():
                summary['cv_skills'][result['filename']][category] = [
                    skill['skill'] for skill in skills
                ]
                
                # Count skills across all CVs
                for skill in skills:
                    skill_name = skill['skill']
                    skill_count[skill_name] = skill_count.get(skill_name, 0) + 1

    # Get top 10 most common skills
    summary['most_common_skills'] = dict(
        sorted(skill_count.items(), key=lambda x: x[1], reverse=True)[:10]
    )

    return summary

def main():
    """Main execution function."""
    logger.info("Starting CV analysis process...")
    
    try:
        # Initialize analyzer
        analyzer = CVAnalyzer(SKILLS_DB)
        
        # Get CV files
        cv_files = get_cv_files()
        if not cv_files:
            logger.error("No CV files found in the specified directory.")
            return
        
        logger.info(f"Found {len(cv_files)} CV files to process.")
        
        # Process CVs
        results = []
        for cv_file in cv_files:
            filename = os.path.basename(cv_file)
            logger.info(f"Processing: {filename}")
            try:
                result = analyzer.analyze_cv(cv_file)
                if result:
                    # Add filename to the result
                    result['filename'] = filename
                    results.append(result)
            except Exception as e:
                logger.error(f"Error processing {filename}: {str(e)}")
        
        # Generate summary
        summary = generate_summary(results)
        results.append({"summary": summary})
        
        # Save results to JSON
        output_file = save_results(results)
        
        # Export skills to Excel
        excel_file = export_skills_to_excel(summary)
        
        # Log summary
        if output_file:
            logger.info("Analysis complete. Summary:")
            logger.info(f"Total CVs processed: {summary['total_cvs_processed']}")
            logger.info(f"CVs with skills found: {summary['cvs_with_skills']}")
            logger.info("Top skills found across all CVs:")
            for skill, count in summary['most_common_skills'].items():
                logger.info(f"  - {skill}: {count}")
        
    except Exception as e:
        logger.error(f"Error in main execution: {str(e)}")
        raise

if __name__ == "__main__":
    main()